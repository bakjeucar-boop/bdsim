#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""app.py — 가설건물 전력 예측 시스템 (Streamlit 버전)"""

import io
import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# 내부 모듈 임포트
from constants import (
    DHW_FACILITY_PARAMS,
    DHW_HEATER_DEFAULTS,
    DHW_HEATER_TYPES,
    ISO18523_PROFILE,
    PRESETS,
    USE_TYPES,
    WATER_SOURCE_PARAMS,
    WEEKEND_MODES,
)
from energy import (
    _fv,
    _iv,
    build_bui,
    calc_electricity_year,
    compute_effective_infil_ach,
    estimate_persons,
    extract_user_period,
    get_restaurant_hvac_times,
    make_restaurant_gains_profile,
    parse_dt,
    run_iso52016_simulation,
)
from weather import (
    OPEN_METEO_FORECAST_DAYS_MAX,
    _safe_replace_year,
    build_epw_year_dataframe,
    calc_supply_temp_series,
    fetch_open_meteo_weather_for_period,
    write_epw_8760_from_open_meteo,
)


# --- 자동 결과 파일 저장 설정 ---
# 나중에 이 기능을 제거하려면 아래 값을 False로 바꾸거나
# save_result_files_to_outputs() 호출부만 삭제하면 됩니다.
AUTO_SAVE_RESULT_FILES = False
AUTO_SAVE_OUTPUT_DIR = (
    Path(__file__).resolve().parents[2] / "outputs"
    if (Path(__file__).resolve().parents[2] / "outputs").exists()
    else Path(__file__).resolve().parent / "outputs"
)
AUTO_SAVE_EXCEL_NAME = "forecast_energy_results.xlsx"
AUTO_SAVE_CSV_NAME = "forecast_site_hourly.csv"
AUTO_SAVE_UNIQUE_FILES = True

# Extra pre-simulation period used only to settle building thermal state.
# Increase for heavy buildings; set to 0 to disable.
SIMULATION_SPINUP_DAYS = 7
WEATHER_PRELOAD_DAYS = 2
OPEN_METEO_ARCHIVE_MIN_DATE = pd.Timestamp("1940-01-01").date()
MAX_SIMULATION_DAYS = 366
DEFAULT_WALL_U = 0.424
DEFAULT_SLAB_U = 0.410
DEFAULT_ROOF_U = 0.428
DEFAULT_DOOR_U = 1.500
DEFAULT_DOOR_AREA_M2 = 6.0
DEFAULT_DOOR_AREA_RATIO_PCT = 0.75
DEFAULT_WINDOW_U = 2.100
DEFAULT_WINDOW_G = 0.583
DEFAULT_COOLING_CAPACITY_KW = 162.4
DEFAULT_COOLING_COP = 3.5
DEFAULT_HEATING_CAPACITY_KW = 182.8
DEFAULT_HEATING_COP = 4.12
DEFAULT_OFFICE_PERSONS = 35
DEFAULT_DHW_BOILER_CAPACITY_KW = 5.9


def default_persons_for_use(area_m2: float, use_type: str) -> int:
    if use_type == "사무실":
        return max(1, int(round(float(area_m2) * DEFAULT_OFFICE_PERSONS / 800.0)))
    if use_type == "식당":
        return default_restaurant_meals(area_m2)
    return estimate_persons(area_m2, use_type)


def default_restaurant_meals(area_m2: float) -> int:
    """식당의 1일 급식 인원 기본값.

    실제 식수 인원이 있으면 사용자가 직접 바꾸는 것이 가장 정확합니다.
    """
    return max(1, int(round(float(area_m2) * 1.0)))


def estimate_kitchen_exhaust_m3h(area_m2: float, meals_per_day: int) -> float:
    """식당 면적과 1일 급식 인원 기반 주방 후드 배기량의 보수적 기본값."""
    estimate = 2.0 * float(area_m2) + 5.0 * float(meals_per_day)
    estimate = max(500.0, min(5000.0, estimate))
    return round(estimate / 10.0) * 10.0


def normalize_dhw_heater_type(value: str) -> str:
    return {
        "전기저항식(저장)": "전기보일러",
        "전기저항식(순간)": "전기보일러",
        "히트펌프": "전기보일러",
        "가스온수기": "가스보일러",
    }.get(value, value if value in DHW_HEATER_TYPES else "전기보일러")


def choose_simulation_year(start_user: pd.Timestamp, end_user: pd.Timestamp) -> int:
    """Choose an EPW simulation year that minimizes calendar remapping artifacts."""
    start = pd.Timestamp(start_user)
    last = pd.Timestamp(end_user) - pd.Timedelta(hours=1)
    if start.year == last.year and not pd.Timestamp(start.year, 12, 31).is_leap_year:
        return int(start.year)
    return 2009

# --- 0. Streamlit 설정 및 최적화 ---
st.set_page_config(
    page_title="Forecast Energy — 가설건물 전력 예측",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)


# Open-Meteo API 호출 캐싱 (UI 재실행 시 중복 요청 방지)
@st.cache_data(ttl=3600)
def cached_fetch_weather(lat: float, lon: float, start_iso: str, end_iso: str):
    return fetch_open_meteo_weather_for_period(
        lat, lon, pd.Timestamp(start_iso), pd.Timestamp(end_iso)
    )


def save_result_files_to_outputs(excel_buf: io.BytesIO, csv_bytes: bytes) -> Dict[str, Path]:
    """Save result files for environments where browser downloads are unclear."""
    AUTO_SAVE_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if AUTO_SAVE_UNIQUE_FILES:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = AUTO_SAVE_OUTPUT_DIR / f"forecast_energy_results_{stamp}.xlsx"
        csv_path = AUTO_SAVE_OUTPUT_DIR / f"forecast_site_hourly_{stamp}.csv"
    else:
        excel_path = AUTO_SAVE_OUTPUT_DIR / AUTO_SAVE_EXCEL_NAME
        csv_path = AUTO_SAVE_OUTPUT_DIR / AUTO_SAVE_CSV_NAME

    excel_path.write_bytes(excel_buf.getvalue())
    csv_path.write_bytes(csv_bytes)

    latest_paths = {}
    for latest_path, source_bytes in [
        (AUTO_SAVE_OUTPUT_DIR / AUTO_SAVE_EXCEL_NAME, excel_buf.getvalue()),
        (AUTO_SAVE_OUTPUT_DIR / AUTO_SAVE_CSV_NAME, csv_bytes),
    ]:
        if latest_path in (excel_path, csv_path):
            continue
        try:
            latest_path.write_bytes(source_bytes)
            latest_paths[latest_path.suffix.lower()] = latest_path
        except PermissionError:
            pass

    return {"excel": excel_path, "csv": csv_path, **latest_paths}


# --- 1. 세션 상태 (Session State) 초기화 ---
if "buildings" not in st.session_state:
    st.session_state.buildings = []
if "last_result" not in st.session_state:
    st.session_state.last_result = None
if "edit_idx" not in st.session_state:
    st.session_state.edit_idx = None

# --- 2. 헬퍼 함수: Excel 다운로드용 Buffer 생성 ---


def generate_excel_buffer(result: Dict[str, Any]) -> io.BytesIO:
    """결과 데이터를 openpyxl을 사용하여 메모리 내 Excel 파일로 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    bldg_results: Dict[str, pd.DataFrame] = result["bldg_results"]
    df_total: pd.DataFrame = result["df_total"]

    HDR_FILL = PatternFill("solid", start_color="1F4E79")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    SUB_FILL = PatternFill("solid", start_color="BDD7EE")
    SUB_FONT = Font(bold=True, name="Arial", size=10)
    BODY_FONT = Font(name="Arial", size=9)
    NUM_FMT = "#,##0.00"
    AL_C = Alignment(horizontal="center")
    AL_R = Alignment(horizontal="right")

    def hdr(ws, row, col, text, fill=HDR_FILL, font=HDR_FONT):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = fill
        c.font = font
        c.alignment = AL_C
        return c

    def num(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=round(float(val), 3))
        c.font = BODY_FONT
        c.number_format = NUM_FMT
        c.alignment = AL_R

    # 시트 1: 통합 시간별
    ws1 = wb.active
    ws1.title = "통합_시간별"
    bldg_names = list(bldg_results.keys())

    hdr(ws1, 1, 1, "날짜·시각")
    ws1.column_dimensions["A"].width = 18
    col = 2
    for nm in bldg_names:
        hdr(ws1, 1, col, nm)
        ws1.column_dimensions[get_column_letter(col)].width = 14
        col += 1
    hdr(ws1, 1, col, "현장 합계 [kWh]")
    ws1.column_dimensions[get_column_letter(col)].width = 16

    row = 2
    for ts, tr in df_total.iterrows():
        ws1.cell(row=row, column=1, value=str(ts)).font = BODY_FONT
        col = 2
        for nm in bldg_names:
            df_b = bldg_results[nm]
            val = (
                float(df_b.loc[ts, "Total_Elec_kWh"])
                if ts in df_b.index
                else 0.0
            )
            num(ws1, row, col, val)
            col += 1
        num(ws1, row, col, float(tr["Total_Elec_kWh"]))
        row += 1

    # 시트 2: 건물별 시간별 상세
    elec_cols = [
        "HVAC_Elec_kWh",
        "Lighting_Elec_kWh",
        "Equip_Elec_kWh",
        "Fan_Elec_kWh",
        "DHW_Elec_kWh",
        "Total_Elec_kWh",
    ]
    sub_cols = ["HVAC", "조명", "콘센트", "팬", "온수", "합계"]
    for nm, df_b in bldg_results.items():
        ws = wb.create_sheet(title=(nm[:28] if len(nm) > 28 else nm))
        hdr(
            ws,
            1,
            1,
            f"{nm} — 시간별 전력량 [kWh]",
            fill=PatternFill("solid", start_color="1F6B75"),
        )
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=7
        )
        ws.column_dimensions["A"].width = 18
        hdr(ws, 2, 1, "날짜·시각", fill=SUB_FILL, font=SUB_FONT)
        for i, col_nm in enumerate(sub_cols, 2):
            hdr(ws, 2, i, col_nm, fill=SUB_FILL, font=SUB_FONT)
            ws.column_dimensions[get_column_letter(i)].width = 12
        row = 3
        for ts, r in df_b.iterrows():
            ws.cell(row=row, column=1, value=str(ts)).font = BODY_FONT
            for i, ec in enumerate(elec_cols, 2):
                num(ws, row, i, r.get(ec, 0.0))
            row += 1

    # 시트 3: 일별 요약
    ws3 = wb.create_sheet(title="일별_요약")
    hdr(ws3, 1, 1, "날짜")
    ws3.column_dimensions["A"].width = 14
    col = 2
    for nm in bldg_names:
        hdr(ws3, 1, col, nm)
        ws3.column_dimensions[get_column_letter(col)].width = 13
        col += 1
    hdr(ws3, 1, col, "현장 합계 [kWh]")
    ws3.column_dimensions[get_column_letter(col)].width = 16

    daily_total = df_total["Total_Elec_kWh"].resample("D").sum()
    row = 2
    for date, total_kwh in daily_total.items():
        ws3.cell(row=row, column=1, value=str(date.date())).font = BODY_FONT
        col = 2
        for nm in bldg_names:
            df_b = bldg_results[nm]
            daily_b = float(
                df_b["Total_Elec_kWh"]
                .resample("D")
                .sum()
                .get(date, 0.0)
            )
            num(ws3, row, col, daily_b)
            col += 1
        num(ws3, row, col, total_kwh)
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --- 3. 사이드바 (현장 공통 설정) ---
with st.sidebar:
    st.header("📍 현장 공통 설정")
    site_lat = st.number_input(
        "위도", value=37.5665, format="%.6f", help="현장의 위도"
    )
    site_lon = st.number_input(
        "경도", value=126.9780, format="%.6f", help="현장의 경도"
    )

    today = pd.Timestamp.now().normalize().date()
    min_select_date = (
        pd.Timestamp(OPEN_METEO_ARCHIVE_MIN_DATE)
        + pd.Timedelta(days=SIMULATION_SPINUP_DAYS + WEATHER_PRELOAD_DAYS)
    ).date()
    max_select_date = (
        pd.Timestamp(today) + pd.Timedelta(days=OPEN_METEO_FORECAST_DAYS_MAX - 1)
    ).date()

    default_start = min(max(today, min_select_date), max_select_date)
    site_start = st.date_input(
        "예측 시작일",
        value=default_start,
        min_value=min_select_date,
        max_value=max_select_date,
        help="Open-Meteo Historical Weather API와 Forecast API가 지원하는 범위 안에서 선택할 수 있습니다.",
    )
    start_user = pd.Timestamp.combine(site_start, pd.Timestamp("00:00").time())

    max_end_by_period = (
        pd.Timestamp(site_start) + pd.Timedelta(days=MAX_SIMULATION_DAYS - 1)
    ).date()
    max_end_date = min(max_select_date, max_end_by_period)
    default_end = min(
        (pd.Timestamp(site_start) + pd.Timedelta(days=6)).date(),
        max_end_date,
    )
    site_end = st.date_input(
        "예측 종료일",
        value=default_end,
        min_value=site_start,
        max_value=max_end_date,
        help="종료일은 계산에 포함됩니다. 현재 엔진 구조상 한 번에 최대 366일까지 계산합니다.",
    )
    end_user = pd.Timestamp.combine(site_end, pd.Timestamp("00:00").time()) + pd.Timedelta(days=1)
    site_days = int((end_user - start_user).days)
    st.caption(
        f"예측 시간은 시작일 00:00부터 종료일 23:00까지입니다. 선택 기간: {site_days}일"
    )

# --- 4. 메인 화면 레이아웃 ---
st.title("📊 Forecast Energy")
st.caption(
    "가설건물 전력 사용량 및 피크 부하 예측 — ISO 52016-1 기반"
)

tab_input, tab_list, tab_result = st.tabs(
    ["🏢 건물 정보 입력", "📋 계산 대상 목록", "📈 시뮬레이션 결과"]
)

# --- 탭 1: 건물 정보 입력 ---
with tab_input:
    st.subheader("새로운 건물 추가 또는 수정")

    # 기본값 설정 (수정 모드 여부에 따라 분기)
    if st.session_state.edit_idx is not None:
        eb = st.session_state.buildings[st.session_state.edit_idx]
    else:
        eb = {}

    # 초기값 세팅을 위한 Use Preset 매핑
    saved_use_type = eb.get("use_type", USE_TYPES[0])
    use_type = st.selectbox(
        "용도",
        options=USE_TYPES,
        index=USE_TYPES.index(saved_use_type) if saved_use_type in USE_TYPES else 0,
    )
    preset = PRESETS[use_type]
    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("##### 📐 기본 건축 정보")
        bldg_name = st.text_input(
            "건물명", value=eb.get("bldg_name", f"건물_{len(st.session_state.buildings)+1}")
        )
        area_m2 = st.number_input(
            "기준면적 [m²]",
            value=float(eb.get("area_m2", 800.0)),
            step=50.0,
        )
        floors = st.number_input(
            "층수", value=int(eb.get("floors", 2)), min_value=1
        )
        floor_h = st.number_input(
            "층고 [m]", value=float(eb.get("floor_h", 3.5)), step=0.1
        )
        wwr = st.number_input(
            "창면적비 (WWR) [-]",
            value=float(eb.get("wwr", 0.20)),
            min_value=0.0,
            max_value=0.95,
            step=0.05,
        )
        azimuth_deg = st.number_input(
            "방위각 [°]", value=float(eb.get("azimuth_deg", 180.0)), step=10.0
        )
        aspect_ratio = st.number_input(
            "장변/단변 비율",
            value=float(eb.get("aspect_ratio", 1.5)),
            min_value=1.0,
            step=0.1,
        )

    with col2:
        st.markdown("##### ⚙️ 냉난방 및 운영")
        heat_set = st.number_input(
            "난방 설정 [°C]",
            value=float(eb.get("heat_set", preset.heat_set)),
            step=0.5,
        )
        cool_set = st.number_input(
            "냉방 설정 [°C]",
            value=float(eb.get("cool_set", preset.cool_set)),
            step=0.5,
        )
        cop_c = st.number_input(
            "냉방기 COP",
            value=float(eb.get("cop_c", DEFAULT_COOLING_COP)),
            min_value=0.1,
            step=0.1,
        )
        cop_h = st.number_input(
            "난방기 COP",
            value=float(eb.get("cop_h", DEFAULT_HEATING_COP)),
            min_value=0.1,
            step=0.1,
        )
        heating_capacity_kw = st.number_input(
            "난방기 용량 [kW]",
            value=float(eb.get("heating_capacity_kw", DEFAULT_HEATING_CAPACITY_KW)),
            min_value=0.1,
            step=1.0,
            help="히트펌프가 공급할 수 있는 최대 난방 열출력입니다.",
        )
        cooling_capacity_kw = st.number_input(
            "냉방기 용량 [kW]",
            value=float(eb.get("cooling_capacity_kw", DEFAULT_COOLING_CAPACITY_KW)),
            min_value=0.1,
            step=1.0,
            help="히트펌프가 공급할 수 있는 최대 냉방 열출력입니다.",
        )

        if use_type == "식당":
            meal_bfst = st.checkbox(
                "조식 제공 (05~09h)", value=bool(eb.get("meal_bfst", preset.meal_bfst))
            )
            meal_lunch = st.checkbox(
                "중식 제공 (10~14h)", value=bool(eb.get("meal_lunch", preset.meal_lunch))
            )
            meal_dinner = st.checkbox(
                "석식 제공 (16~21h)", value=bool(eb.get("meal_dinner", preset.meal_dinner))
            )
            auto_hvac_start, auto_hvac_end = get_restaurant_hvac_times(
                meal_bfst, meal_lunch, meal_dinner
            )
        else:
            meal_bfst, meal_lunch, meal_dinner = False, False, False
            auto_hvac_start = int(eb.get("hvac_start", preset.hvac_start))
            auto_hvac_end = int(eb.get("hvac_end", preset.hvac_end))

        hvac_start = st.number_input(
            "HVAC 시작 [h]",
            value=auto_hvac_start,
            min_value=0,
            max_value=23,
            disabled=(use_type == "식당"),
        )
        hvac_end = st.number_input(
            "HVAC 종료 [h]",
            value=auto_hvac_end,
            min_value=0,
            max_value=23,
            disabled=(use_type == "식당"),
        )
        saved_weekend_mode = eb.get("weekend_mode")
        if saved_weekend_mode not in WEEKEND_MODES:
            saved_sat = eb.get("sat_mode", preset.sat_mode)
            saved_sun = eb.get("sun_mode", preset.sun_mode)
            saved_weekend_mode = saved_sat if saved_sat == saved_sun else preset.sun_mode
        if saved_weekend_mode not in WEEKEND_MODES:
            saved_weekend_mode = "없음(OFF)"
        weekend_mode = st.selectbox(
            "주말 운영",
            options=WEEKEND_MODES,
            index=WEEKEND_MODES.index(saved_weekend_mode),
            help="pybuildingenergy는 토요일과 일요일을 하나의 주말 프로파일로 계산하므로 동일한 주말 운영 조건을 적용합니다.",
        )
        sat_mode = weekend_mode
        sun_mode = weekend_mode

    with col3:
        st.markdown("##### 🚰 온수 및 환기")
        dhw_facility = {
            "사무실": "세면",
            "작업장": "세면",
            "숙소": "샤워",
            "식당": "주방",
        }.get(use_type, "세면")
        dhw_heater_type = st.selectbox(
            "온수기 종류",
            options=DHW_HEATER_TYPES,
            index=DHW_HEATER_TYPES.index(
                normalize_dhw_heater_type(eb.get("dhw_heater_type", preset.default_dhw_heater))
            ),
        )

        default_persons = default_persons_for_use(area_m2, use_type)
        dhw_person_label = "1일 급식 인원" if use_type == "식당" else "재실 인원"
        dhw_persons = st.number_input(
            dhw_person_label,
            value=int(eb.get("dhw_persons", default_persons)),
            min_value=1,
            help=(
                "식당 온수부하는 상시 재실자보다 식수 인원 영향이 크므로 1일 급식 인원을 사용합니다."
                if use_type == "식당"
                else None
            ),
        )
        dhw_capacity_kw = st.number_input(
            "온수기 용량 [kW]",
            value=float(eb.get("dhw_capacity_kw", DEFAULT_DHW_BOILER_CAPACITY_KW)),
            min_value=0.0,
            step=0.1,
            help="전기보일러는 정격 소비전력, 가스보일러/외부공급은 공급 가능한 열용량 기준입니다. 0이면 용량 제한을 적용하지 않습니다.",
        )
        water_source_options = list(WATER_SOURCE_PARAMS.keys())
        water_source = st.selectbox(
            "급수 공급원",
            options=water_source_options,
            index=water_source_options.index(
                eb.get("water_source", "지중매설 (1m+)")
                if eb.get("water_source", "지중매설 (1m+)") in water_source_options
                else "지중매설 (1m+)"
            ),
        )

        mechanical_vent = st.checkbox(
            "기계환기 있음",
            value=bool(eb.get("mechanical_vent", float(eb.get("oa_m3h", preset.oa_m3h)) > 0.0)),
        )
        if mechanical_vent:
            oa_m3h = st.number_input(
                "외기공급량 [m³/h]", value=float(eb.get("oa_m3h", preset.oa_m3h))
            )
        else:
            oa_m3h = 0.0

        # 식당 전용 후드 배기 설정
        if use_type == "식당":
            kitchen_exh_auto = st.checkbox(
                "주방 후드 배기 자동추정",
                value=bool(eb.get("kitchen_exh_auto", True)),
            )
            estimated_kitchen_exh = estimate_kitchen_exhaust_m3h(area_m2, dhw_persons)
            if kitchen_exh_auto:
                kitchen_exh = estimated_kitchen_exh
                st.caption(f"자동추정값: {kitchen_exh:,.0f} m³/h")
            else:
                kitchen_exh = st.number_input(
                    "주방 후드 배기 [m³/h]",
                    value=float(
                        eb.get("kitchen_exh", estimated_kitchen_exh)
                    ),
                    min_value=0.0,
                    step=50.0,
                )
        else:
            kitchen_exh_auto = False
            kitchen_exh = 0.0

    # 고급 설정: 접이식 패널 (Expanders)
    with st.expander("🛠️ 고급 설정 (부하 밀도 및 외피 정보)"):
        ae_col1, ae_col2 = st.columns(2)
        with ae_col1:
            gain_heat = st.number_input(
                "인체+기기 발열 [W/m²]",
                value=float(
                    eb.get("gain_heat", preset.internal_gain_heat_wm2)
                ),
            )
            lighting_wm2 = st.number_input(
                "조명 밀도 [W/m²]",
                value=float(eb.get("lighting_wm2", preset.lighting_wm2)),
            )
            equip_elec_wm2 = st.number_input(
                "콘센트·가전 [W/m²]",
                value=float(
                    eb.get("equip_elec_wm2", preset.equip_elec_wm2)
                ),
            )
            infil_base = st.number_input(
                "기본 침기율 [ACH]",
                value=float(eb.get("infil_base", preset.base_infil_ach)),
            )
        with ae_col2:
            roof_u = st.number_input(
                "지붕 열관류율 [W/m²K]", value=float(eb.get("roof_u", DEFAULT_ROOF_U))
            )
            wall_u = st.number_input(
                "외벽 열관류율 [W/m²K]", value=float(eb.get("wall_u", DEFAULT_WALL_U))
            )
            slab_u = st.number_input(
                "바닥 열관류율 [W/m²K]", value=float(eb.get("slab_u", DEFAULT_SLAB_U))
            )
            door_u = st.number_input(
                "문 열관류율 [W/m²K]", value=float(eb.get("door_u", DEFAULT_DOOR_U))
            )
            default_door_ratio = float(
                eb.get(
                    "door_area_ratio_pct",
                    100.0 * float(eb.get("door_area_m2", DEFAULT_DOOR_AREA_M2)) / max(float(area_m2), 1.0),
                )
            )
            door_area_ratio_pct = st.number_input(
                "문 면적 비율 [%]",
                value=default_door_ratio,
                min_value=0.0,
                step=0.1,
                help="바닥면적 대비 문 면적 비율입니다. 기본 0.75%는 800 m² 기준 약 6 m²입니다.",
            )
            door_area_m2 = float(area_m2) * float(door_area_ratio_pct) / 100.0
            win_u = st.number_input(
                "창호 열관류율 [W/m²K]", value=float(eb.get("win_u", DEFAULT_WINDOW_U))
            )
            win_g = st.number_input(
                "창호 SHGC (g-value) [-]",
                value=float(eb.get("win_g", DEFAULT_WINDOW_G)),
            )

    # 데이터 수집 및 등록
    building_data = {
        "bldg_name": bldg_name,
        "use_type": use_type,
        "area_m2": area_m2,
        "floors": floors,
        "floor_h": floor_h,
        "wwr": wwr,
        "azimuth_deg": azimuth_deg,
        "aspect_ratio": aspect_ratio,
        "heat_set": heat_set,
        "cool_set": cool_set,
        "hvac_start": hvac_start,
        "hvac_end": hvac_end,
        "sat_mode": sat_mode,
        "sun_mode": sun_mode,
        "weekend_mode": weekend_mode,
        "dhw_facility": dhw_facility,
        "dhw_heater_type": dhw_heater_type,
        "dhw_persons": dhw_persons,
        "dhw_capacity_kw": dhw_capacity_kw,
        "water_source": water_source,
        "mechanical_vent": mechanical_vent,
        "oa_m3h": oa_m3h,
        "kitchen_exh": kitchen_exh,
        "kitchen_exh_auto": kitchen_exh_auto,
        "meal_bfst": meal_bfst,
        "meal_lunch": meal_lunch,
        "meal_dinner": meal_dinner,
        "gain_heat": gain_heat,
        "lighting_wm2": lighting_wm2,
        "equip_elec_wm2": equip_elec_wm2,
        "infil_base": infil_base,
        "roof_u": roof_u,
        "wall_u": wall_u,
        "slab_u": slab_u,
        "door_u": door_u,
        "door_area_ratio_pct": door_area_ratio_pct,
        "door_area_m2": door_area_m2,
        "win_u": win_u,
        "win_g": win_g,
        # 기본 fallback 값 처리
        "gains_start": preset.gains_start,
        "gains_end": preset.gains_end,
        "cop_h": cop_h,
        "cop_c": cop_c,
        "heating_capacity_kw": heating_capacity_kw,
        "cooling_capacity_kw": cooling_capacity_kw,
        "vent_start": 8,
        "vent_end": 18,
        "fan_sp": 0.0,
        "dhw_cop": DHW_HEATER_DEFAULTS.get(dhw_heater_type, {}).get("cop", 0.0),
        "dhw_t_hot_shower": DHW_HEATER_DEFAULTS.get(dhw_heater_type, {}).get(
            "t_hot_shower", 45.0
        ),
        "dhw_t_hot_kitchen": DHW_HEATER_DEFAULTS.get(dhw_heater_type, {}).get(
            "t_hot_kitchen", 80.0
        ),
    }

    # 버튼 UI
    if st.session_state.edit_idx is not None:
        if st.button("✏️ 수정 완료", type="primary"):
            st.session_state.buildings[st.session_state.edit_idx] = building_data
            st.session_state.edit_idx = None
            st.success(f"[{bldg_name}] 수정되었습니다.")
            st.rerun()
    else:
        if st.button("➕ 목록에 추가", type="primary"):
            # 이름 중복 방지
            existing_names = [b["bldg_name"] for b in st.session_state.buildings]
            if bldg_name in existing_names:
                building_data["bldg_name"] = f"{bldg_name}_{len(st.session_state.buildings)+1}"

            st.session_state.buildings.append(building_data)
            st.success(f"[{building_data['bldg_name']}] 목록에 추가되었습니다.")
            st.rerun()

# --- 탭 2: 계산 대상 목록 ---
with tab_list:
    st.subheader("등록된 건물 목록")

    if not st.session_state.buildings:
        st.info("등록된 건물이 없습니다. [건물 정보 입력] 탭에서 건물을 추가해 주세요.")
    else:
        # 데이터프레임 뷰로 전환
        bldg_summary = []
        for i, b in enumerate(st.session_state.buildings):
            bldg_summary.append(
                {
                    "No": i + 1,
                    "건물명": b["bldg_name"],
                    "용도": b["use_type"],
                    "면적 [m²]": b["area_m2"],
                    "층수": b["floors"],
                    "주말운영": b.get("weekend_mode", b.get("sat_mode", "")),
                }
            )

        df_summary = pd.DataFrame(bldg_summary)
        st.dataframe(df_summary, use_container_width=True, hide_index=True)

        col_action1, col_action2, col_action3 = st.columns([1, 1, 4])
        with col_action1:
            mod_idx = st.number_input(
                "수정할 번호",
                min_value=1,
                max_value=len(st.session_state.buildings),
                step=1,
            )
            if st.button("📝 선택 수정"):
                st.session_state.edit_idx = mod_idx - 1
                st.success(
                    f"{mod_idx}번 건물을 수정 모드로 전환했습니다. '건물 정보 입력' 탭을 확인하세요."
                )
                st.rerun()

        with col_action2:
            del_idx = st.number_input(
                "삭제할 번호",
                min_value=1,
                max_value=len(st.session_state.buildings),
                step=1,
            )
            if st.button("🗑️ 선택 삭제", type="secondary"):
                del_name = st.session_state.buildings[del_idx - 1]["bldg_name"]
                st.session_state.buildings.pop(del_idx - 1)
                st.success(f"{del_name} 건물이 삭제되었습니다.")
                st.rerun()

        st.divider()

        # 시뮬레이션 가동 버튼
        if st.button("🚀 전체 시뮬레이션 계산 실행", type="primary", use_container_width=True):
            buildings = st.session_state.buildings
            epw_year = choose_simulation_year(start_user, end_user)

            # 7. Core Logic - 뷰 및 진행 상황 바
            with st.status("물리 엔진 시뮬레이션 진행 중...", expanded=True) as status:
                try:
                    spinup_days = max(0, int(SIMULATION_SPINUP_DAYS))
                    sim_extract_start = start_user - pd.Timedelta(days=spinup_days)
                    sim_extract_days = int(site_days) + spinup_days
                    weather_start = sim_extract_start - pd.Timedelta(days=WEATHER_PRELOAD_DAYS)
                    weather_end = end_user

                    # Step A: 기상 데이터 수집
                    status.write("📡 Step A: Open-Meteo 기상 수집 중...")
                    df_open, wmeta = cached_fetch_weather(
                        site_lat,
                        site_lon,
                        weather_start.isoformat(),
                        weather_end.isoformat(),
                    )

                    # Step C: EPW 생성
                    status.write("💾 Step C: EPW 기상 파일 생성 중...")
                    base_8760, df_mapped = build_epw_year_dataframe(
                        df_open, wmeta, epw_year
                    )

                    ws_sim = _safe_replace_year(weather_start, epw_year)
                    we_sim = _safe_replace_year(end_user, epw_year)

                    tmp_dir = tempfile.mkdtemp()
                    epw_path = os.path.join(tmp_dir, f"om_{epw_year}.epw")
                    write_epw_8760_from_open_meteo(
                        base_8760,
                        wmeta,
                        epw_path,
                        compensate_pybui_tz_roll=True,
                    )

                    # Step D: 건물별 시뮬레이션
                    bldg_results = {}
                    n_bldgs = len(buildings)

                    for idx, bp in enumerate(buildings):
                        nm = bp["bldg_name"]
                        ut = bp["use_type"]
                        status.write(
                            f"🏢 Step D-{idx+1}/{n_bldgs}: [{nm}] ISO 52016 계산 중..."
                        )

                        area_m2 = bp["area_m2"]
                        water_source_b = bp.get("water_source", "지중매설 (1m+)")
                        t_supply, freeze_warns = calc_supply_temp_series(
                            df_open, water_source_b
                        )
                        dhw_persons = bp["dhw_persons"]
                        if dhw_persons <= 0:
                            dhw_persons = default_persons_for_use(area_m2, ut)
                        occ_ratio = min(
                            1.0, dhw_persons / max(1, estimate_persons(area_m2, ut))
                        )

                        # 식당 특화 처리
                        restaurant_gains_prof = None
                        hvac_start, hvac_end = bp["hvac_start"], bp["hvac_end"]
                        gains_start, gains_end = (
                            bp["gains_start"],
                            bp["gains_end"],
                        )

                        if ut == "식당":
                            restaurant_gains_prof = make_restaurant_gains_profile(
                                bp["meal_bfst"], bp["meal_lunch"], bp["meal_dinner"]
                            )
                            hvac_start, hvac_end = get_restaurant_hvac_times(
                                bp["meal_bfst"], bp["meal_lunch"], bp["meal_dinner"]
                            )
                            gains_start = min(hvac_start + 1, 23)
                            gains_end = max(hvac_end - 1, 0)

                        # 풍속 보정
                        infil_eff, v75 = compute_effective_infil_ach(
                            area_m2,
                            bp["floor_h"],
                            bp["infil_base"],
                            bp["oa_m3h"],
                            bp["kitchen_exh"],
                            bp["vent_start"],
                            bp["vent_end"],
                            df_open=df_open,
                            start_user=start_user,
                            days=site_days,
                        )

                        bui = build_bui(
                            name=f"{nm}_{ut}",
                            lat=site_lat,
                            lon=site_lon,
                            area_m2=area_m2,
                            n_floors=bp["floors"],
                            floor_height_m=bp["floor_h"],
                            wwr=bp["wwr"],
                            azimuth_deg=bp["azimuth_deg"],
                            aspect_ratio=bp["aspect_ratio"],
                            roof_u=bp["roof_u"],
                            wall_u=bp["wall_u"],
                            slab_u=bp["slab_u"],
                            door_u=bp.get("door_u", DEFAULT_DOOR_U),
                            door_area_m2=bp.get("door_area_m2", DEFAULT_DOOR_AREA_M2),
                            win_u=bp["win_u"],
                            win_g=bp["win_g"],
                            heat_set=bp["heat_set"],
                            cool_set=bp["cool_set"],
                            hvac_start=hvac_start,
                            hvac_end=hvac_end,
                            heating_capacity_kw=bp.get(
                                "heating_capacity_kw", DEFAULT_HEATING_CAPACITY_KW
                            ),
                            cooling_capacity_kw=bp.get(
                                "cooling_capacity_kw", DEFAULT_COOLING_CAPACITY_KW
                            ),
                            gains_start=gains_start,
                            gains_end=gains_end,
                            internal_gain_heat_wm2=bp["gain_heat"],
                            lighting_wm2=bp["lighting_wm2"],
                            infil_ach=infil_eff,
                            sat_mode=bp["sat_mode"],
                            sun_mode=bp["sun_mode"],
                            occupancy_ratio=occ_ratio,
                            restaurant_gains_prof=restaurant_gains_prof,
                        )

                        hs_year = run_iso52016_simulation(
                            bui, epw_path, epw_year
                        )
                        sim_year = int(hs_year.index.min().year)

                        fp = DHW_FACILITY_PARAMS.get(
                            bp["dhw_facility"],
                            DHW_FACILITY_PARAMS["없음"],
                        )
                        df_year = calc_electricity_year(
                            hs_year,
                            cop_h=bp["cop_h"],
                            cop_c=bp["cop_c"],
                            area_m2=area_m2,
                            lighting_wm2=bp["lighting_wm2"],
                            equip_elec_wm2=bp["equip_elec_wm2"],
                            gains_start=gains_start,
                            gains_end=gains_end,
                            supply_oa_m3h=bp["oa_m3h"],
                            kitchen_exh_m3h=bp["kitchen_exh"],
                            vent_start=bp["vent_start"],
                            vent_end=bp["vent_end"],
                            fan_sp=bp["fan_sp"],
                            t_supply=t_supply,
                            facility_type=bp["dhw_facility"],
                            heater_type=normalize_dhw_heater_type(bp["dhw_heater_type"]),
                            persons=dhw_persons,
                            shower_lpd=fp["shower_lpd"],
                            shower_t_hot=bp["dhw_t_hot_shower"],
                            kitchen_lpd=fp["kitchen_lpd"],
                            kitchen_t_hot=bp["dhw_t_hot_kitchen"],
                            dhw_cop=bp["dhw_cop"],
                            dhw_capacity_kw=bp.get("dhw_capacity_kw", DEFAULT_DHW_BOILER_CAPACITY_KW),
                            use_type=ut,
                            occupancy_ratio=occ_ratio,
                            df_open=df_open,
                            sat_mode=bp["sat_mode"],
                            sun_mode=bp["sun_mode"],
                            restaurant_gains_prof=restaurant_gains_prof,
                        )

                        df_spin = extract_user_period(
                            df_year,
                            df_open,
                            sim_extract_start,
                            sim_extract_days,
                            sim_year,
                        )
                        df_out = df_spin.loc[
                            (df_spin.index >= start_user) & (df_spin.index < end_user)
                        ].copy()
                        bldg_results[nm] = df_out

                    # 임시 파일 정리
                    try:
                        os.remove(epw_path)
                        os.rmdir(tmp_dir)
                    except:
                        pass

                    # 결과 병합
                    df_total = (
                        pd.concat(
                            [
                                df[
                                    [
                                        "Total_Elec_kWh",
                                        "HVAC_Elec_kWh",
                                        "Lighting_Elec_kWh",
                                        "Equip_Elec_kWh",
                                        "Fan_Elec_kWh",
                                        "DHW_Elec_kWh",
                                    ]
                                ]
                                for df in bldg_results.values()
                            ],
                            axis=1,
                            keys=bldg_results.keys(),
                        )
                        .groupby(level=1, axis=1)
                        .sum()
                    )

                    meta_water_source = buildings[0].get("water_source", "지중매설 (1m+)") if buildings else "지중매설 (1m+)"
                    t_supply_meta, _ = calc_supply_temp_series(df_open, meta_water_source)
                    sp_param = WATER_SOURCE_PARAMS.get(meta_water_source, {})
                    df_note = ""
                    if sp_param.get("soil_col") is not None:
                        df_note = f' (depth_factor={sp_param.get("depth_factor",1.0)}, 54cm 기반)'

                    t_s_p = t_supply_meta.reindex(df_total.index, method="nearest")

                    st.session_state.last_result = {
                        "bldg_results": bldg_results,
                        "df_total": df_total,
                        "meta": {
                            "start": start_user,
                            "end": end_user,
                            "timezone": wmeta.get("timezone", ""),
                            "weather_source": wmeta.get("source", ""),
                            "weather_start": wmeta.get("weather_start", ""),
                            "weather_end": wmeta.get("weather_end", ""),
                            "spinup_days": spinup_days,
                            "water_source": meta_water_source,
                            "t_sup_mean": float(t_s_p.mean()),
                            "t_sup_min": float(t_s_p.min()),
                            "t_sup_max": float(t_s_p.max()),
                            "df_note": df_note,
                            "n_bldgs": len(buildings),
                        },
                    }

                    status.update(
                        label="✅ 시뮬레이션 계산 완료!", state="complete"
                    )
                    st.toast("모든 건물의 시뮬레이션이 성공적으로 끝났습니다.")

                except Exception as e:
                    status.update(
                        label="❌ 시뮬레이션 실패", state="error"
                    )
                    st.error(f"오류가 발생했습니다: {str(e)}")
                    st.exception(e)


# --- 탭 3: 계산 결과 ---
with tab_result:
    if st.session_state.last_result is None:
        st.info("시뮬레이션을 실행하면 결과가 여기에 표시됩니다.")
    else:
        res = st.session_state.last_result
        meta = res["meta"]
        df_total = res["df_total"]
        bldg_results = res["bldg_results"]

        # KPI 메트릭 카드 상단 배치
        st.subheader("🏁 종합 및 현장 합계 요약")
        kpi1, kpi2, kpi3, kpi4 = st.columns(4)
        total_site_kwh = df_total["Total_Elec_kWh"].sum()
        peak_site_kw = df_total["Total_Elec_kWh"].max()
        peak_site_time = df_total["Total_Elec_kWh"].idxmax()

        kpi1.metric(
            label="현장 총 소비 전력량", value=f"{total_site_kwh:,.1f} kWh"
        )
        kpi2.metric(label="현장 피크 전력 수요", value=f"{peak_site_kw:,.2f} kW")
        kpi3.metric(
            label="현장 피크 시간", value=peak_site_time.strftime("%m-%d %H:%M")
        )
        kpi4.metric(label="급수 평균 온도", value=f"{meta['t_sup_mean']:.1f} °C")

        st.divider()

        # 시각화 영역
        st.subheader("📈 예측기간 시간별 전력수요")
        fig_line = go.Figure()
        fig_line.add_trace(
            go.Scatter(
                x=df_total.index,
                y=df_total["Total_Elec_kWh"],
                mode="lines",
                name="현장 총 전력",
                fill="tozeroy",
                line=dict(color="#1F4E79", width=3),
            )
        )
        # 주요 용도별 데이터 추가
        for sub, label in [
            ("HVAC_Elec_kWh", "냉난방"),
            ("Lighting_Elec_kWh", "조명"),
            ("Equip_Elec_kWh", "콘센트"),
            ("DHW_Elec_kWh", "온수"),
        ]:
            fig_line.add_trace(
                go.Scatter(
                    x=df_total.index,
                    y=df_total[sub],
                    mode="lines",
                    name=label,
                    line=dict(width=1.5, dash="dot"),
                )
            )

        fig_line.update_layout(
            xaxis_title="날짜 및 시각",
            yaxis_title="전력수요 [kW]",
            hovermode="x unified",
            height=480,
            margin=dict(l=20, r=20, t=20, b=20),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        )
        st.plotly_chart(fig_line, use_container_width=True)

        st.divider()

        st.subheader("🗓️ 예측 날짜별 시간대 총 전력 사용량")
        heatmap_df = df_total[["Total_Elec_kWh"]].copy()
        heatmap_df["날짜"] = heatmap_df.index.strftime("%Y-%m-%d")
        heatmap_df["시간"] = heatmap_df.index.hour
        heatmap_table = heatmap_df.pivot_table(
            index="날짜",
            columns="시간",
            values="Total_Elec_kWh",
            aggfunc="sum",
        ).reindex(columns=range(24))
        heatmap_table.columns = [f"{h:02d}시" for h in heatmap_table.columns]
        st.dataframe(
            heatmap_table.style.format("{:.1f}").background_gradient(
                cmap="Reds", axis=None
            ),
            use_container_width=True,
            height=min(420, 74 + 35 * len(heatmap_table)),
        )

        st.divider()

        st.subheader("🏢 건물별 세부 요약")
        # 건물별 세부 요약 테이블
        bldg_det = []
        for nm, df_b in bldg_results.items():
            bldg_det.append(
                {
                    "건물명": nm,
                    "총 전력량 [kWh]": round(df_b["Total_Elec_kWh"].sum(), 1),
                    "냉난방 (HVAC)": round(df_b["HVAC_Elec_kWh"].sum(), 1),
                    "조명": round(df_b["Lighting_Elec_kWh"].sum(), 1),
                    "콘센트": round(df_b["Equip_Elec_kWh"].sum(), 1),
                    "팬": round(df_b["Fan_Elec_kWh"].sum(), 1),
                    "급탕 (DHW)": round(df_b["DHW_Elec_kWh"].sum(), 1),
                    "피크 부하 [kW]": round(df_b["Total_Power_kW"].max(), 2),
                    "피크 시간": df_b["Total_Power_kW"].idxmax().strftime("%m-%d %H:%M"),
                }
            )
        bldg_det.append(
            {
                "건물명": "합계",
                "총 전력량 [kWh]": round(df_total["Total_Elec_kWh"].sum(), 1),
                "냉난방 (HVAC)": round(df_total["HVAC_Elec_kWh"].sum(), 1),
                "조명": round(df_total["Lighting_Elec_kWh"].sum(), 1),
                "콘센트": round(df_total["Equip_Elec_kWh"].sum(), 1),
                "팬": round(df_total["Fan_Elec_kWh"].sum(), 1),
                "급탕 (DHW)": round(df_total["DHW_Elec_kWh"].sum(), 1),
                "피크 부하 [kW]": round(df_total["Total_Elec_kWh"].max(), 2),
                "피크 시간": df_total["Total_Elec_kWh"].idxmax().strftime("%m-%d %H:%M"),
            }
        )
        st.dataframe(
            pd.DataFrame(bldg_det), use_container_width=True, hide_index=True
        )

        st.divider()

        # 다운로드 영역
        st.subheader("📥 데이터 다운로드")
        with st.spinner("결과 파일을 준비 중입니다..."):
            excel_buf = generate_excel_buffer(res)
            csv_site = df_total.to_csv().encode("utf-8-sig")

        if AUTO_SAVE_RESULT_FILES:
            saved_paths = save_result_files_to_outputs(excel_buf, csv_site)
            st.success(
                "결과 파일이 outputs 폴더에 새 파일명으로 자동 저장되었습니다. "
                f"Excel: {saved_paths['excel']} / CSV: {saved_paths['csv']}"
            )

        col_dw1, col_dw2 = st.columns(2)

        with col_dw1:
            st.download_button(
                label="📥 통합 결과 Excel 파일 다운로드",
                data=excel_buf.getvalue(),
                file_name=AUTO_SAVE_EXCEL_NAME,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with col_dw2:
            st.download_button(
                label="📥 현장 합계 CSV 파일 다운로드",
                data=csv_site,
                file_name=AUTO_SAVE_CSV_NAME,
                mime="text/csv",
                use_container_width=True,
            )




